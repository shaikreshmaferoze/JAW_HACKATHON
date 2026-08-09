from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.parsing.normalization import clean_text


def _json_value(value: object) -> object:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def extract_workbook(path: str | Path) -> dict[str, Any]:
    path = Path(path)
    wb_formula = load_workbook(path, data_only=False, read_only=False)
    wb_values = load_workbook(path, data_only=True, read_only=True)
    sheets: list[dict[str, Any]] = []
    text_parts: list[str] = []

    for ws in wb_formula.worksheets:
        ws_values = wb_values[ws.title]
        rows: list[list[object]] = []
        formulas: list[dict[str, str]] = []
        for row in ws.iter_rows():
            values = []
            for cell in row:
                values.append(_json_value(cell.value))
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append({"cell": cell.coordinate, "formula": cell.value})
            rows.append(values)

        value_rows: list[list[object]] = []
        for row in ws_values.iter_rows(values_only=True):
            value_rows.append([_json_value(v) for v in row])

        lines = [f"Sheet: {ws.title} ({ws.max_row}x{ws.max_column})"]
        for row in value_rows:
            cells = [clean_text(v) for v in row if v is not None and clean_text(v)]
            if cells:
                lines.append(" | ".join(cells))
        text_parts.extend(lines)
        sheets.append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "rows": rows,
                "values": value_rows,
                "formulas": formulas,
            }
        )

    text = "\n".join(text_parts)
    return {
        "pages": None,
        "sheets": sheets,
        "sheet_names": [s["name"] for s in sheets],
        "text": clean_text(text),
        "extraction_method": "openpyxl",
        "extraction_quality": 1.0,
        "char_count": len(clean_text(text)),
    }

